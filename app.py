import base64
import datetime
import functools
import os
import sqlite3
import uuid
import requests
import logging
import io
import resend

from dotenv import load_dotenv
from flask import Flask, jsonify, redirect, render_template, request, session, url_for, make_response, flash
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from spin import spin_bp

load_dotenv()

logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)

# Fix proxied HTTPS on Render
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

app.secret_key = os.getenv("SECRET_KEY", "dev_secret_key_change_in_production")
DB_PATH = os.getenv("DB_PATH", "app_database.db")
app.register_blueprint(spin_bp)

app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = bool(os.getenv("RENDER"))
app.config['SESSION_COOKIE_HTTPONLY'] = True

if os.getenv("RENDER"):
    app.config["PREFERRED_URL_SCHEME"] = "https"

# ----------------------------------------------------------------
# RESEND EMAIL CONFIG
# ----------------------------------------------------------------
resend.api_key = os.getenv("RESEND_API_KEY")

# ----------------------------------------------------------------
# TOKEN SERIALIZER
# ----------------------------------------------------------------
serializer = URLSafeTimedSerializer(app.secret_key)

# ----------------------------------------------------------------
# PAYHERO CONFIG
# ----------------------------------------------------------------
PAYHERO_BASE_URL   = os.getenv("PAYHERO_BASE_URL", "https://backend.payhero.co.ke/api/v2")
PAYHERO_CHANNEL_ID = os.getenv("PAYHERO_CHANNEL_ID", "6532")
PAYHERO_PROVIDER   = os.getenv("PAYHERO_PROVIDER", "m-pesa")

if os.getenv("RENDER"):
    CALLBACK_URL = "https://gainpesa-zatz.onrender.com/callback"
else:
    CALLBACK_URL = os.getenv("CALLBACK_URL", "https://cedrick-subdiscoid-drake.ngrok-free.de/callback")

API_USERNAME = os.getenv("API_USERNAME")
API_PASSWORD = os.getenv("API_PASSWORD", "gMMRAHjO3snOZgQI7kS2xPpLlXLcylaKqaW5CJXd")

ACTIVATION_FEE         = 100.0
USD_TO_KES             = 130.0
MIN_BINARY_DEPOSIT_KES = round(1.0 * USD_TO_KES, 2)


def get_auth_header():
    auth = f"{API_USERNAME}:{API_PASSWORD}"
    return f"Basic {base64.b64encode(auth.encode()).decode()}"


# ----------------------------------------------------------------
# DATABASE
# ----------------------------------------------------------------
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            email TEXT PRIMARY KEY, username TEXT UNIQUE, password_hash TEXT, phone TEXT,
            balance REAL DEFAULT 0.0, spin_balance REAL DEFAULT 0.0,
            binary_balance REAL DEFAULT 0.0, binary_deposited REAL DEFAULT 0.0,
            binary_winnings REAL DEFAULT 0.0, total_earned REAL DEFAULT 0.0,
            total_withdrawn REAL DEFAULT 0.0, total_referred INTEGER DEFAULT 0,
            is_active BOOLEAN DEFAULT 0, referral_code TEXT UNIQUE,
            referred_by TEXT, joined_at TEXT, reset_token TEXT, token_expiry TEXT
        )
    """)
    for col, typedef in {
        "spin_balance": "REAL DEFAULT 0.0", "binary_balance": "REAL DEFAULT 0.0",
        "binary_deposited": "REAL DEFAULT 0.0", "binary_winnings": "REAL DEFAULT 0.0",
        "reset_token": "TEXT", "token_expiry": "TEXT",
    }.items():
        try:
            conn.execute(f"ALTER TABLE users ADD COLUMN {col} {typedef}")
        except Exception:
            pass

    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            ext_ref TEXT PRIMARY KEY, email TEXT, type TEXT DEFAULT 'activation',
            status TEXT, amount REAL DEFAULT 0.0, FOREIGN KEY(email) REFERENCES users(email)
        )
    """)
    for col, td in [("type", "TEXT DEFAULT 'activation'"), ("amount", "REAL DEFAULT 0.0")]:
        try:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {td}")
        except Exception:
            pass

    conn.execute("""
        CREATE TABLE IF NOT EXISTS withdrawals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT, amount REAL,
            mpesa_number TEXT, status TEXT, date TEXT, FOREIGN KEY(email) REFERENCES users(email)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS binary_trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT, asset TEXT, amount REAL,
            direction TEXT, status TEXT, payout REAL, timestamp TEXT,
            FOREIGN KEY(email) REFERENCES users(email)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS admin_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, admin_username TEXT,
            target_email TEXT, action_type TEXT, amount REAL, timestamp TEXT
        )
    """)
    conn.commit()
    conn.close()


init_db()


def seed_active_users():
    # (username, email, is_active, wallet_balance, binary_balance)
    rows = [
        ("Offmask","titusilla611@gmail.com",0,0.00,0.00),
        ("Anonymous","agentheroqash@gmail.com",0,0.00,0.00),
        ("Stormin","thuojustus39@gmail.com",0,0.00,0.00),
        ("Jinwoo","narutuuzumaki307@gmail.com",0,0.00,0.00),
        ("Delmartin Junior","delmartinouma471@gmail.com",0,0.00,0.00),
        ("Gopato","mukhwanapatrick73@gmail.com",1,0.00,100.00),
        ("Denisah","denisdenisah777@gmail.com",0,0.00,0.00),
        ("Limo","christopherruto208@gmail.com",1,0.00,120.00),
        ("NJUGUNA MBURU","njugunamburu2001@gmail.com",0,0.00,0.00),
        ("Trader","lewish1876@gmail.com",0,0.00,0.00),
        ("vigimath","litwakamathew@gmail.com",0,0.00,0.00),
        ("mathewlitwaka","mathewlitwaka@gmail.com",1,0.00,100.00),
        ("Isaac letwati","letwatiisaac@gmail.com",0,0.00,0.00),
        ("Judie20","judiekeylah@gmail.com",1,0.00,30.00),
        ("Manuu","manuukorir534@gmail.com",1,0.00,40.00),
        ("Kip","devinkiprop224@gmail.com",0,0.00,0.00),
        ("Alvin Wanyonyi","alvinwanyonyi21@gmail.com",0,0.00,0.00),
        ("Mkurungenzi wuapari","saoliedward@gmail.com",0,0.00,0.00),
        ("Yasporino","yaspicharlie@gmail.com",0,0.00,0.00),
        ("Rono@","kipkoechenock623@gmail.com",0,0.00,0.00),
        ("Ndai","ericndai8@gmail.com",0,0.00,0.00),
        ("WNCVVQ","shadrackkiprono732@gmail.com",0,0.00,0.00),
        ("josefx","jfx7695@gmail.com",1,0.00,100.00),
        ("Luvai","dennisluvai14@gmail.com",0,0.00,0.00),
        ("Chumba","chumbamisael@gmail.com",1,0.00,40.00),
        ("Joshua","joshuamathuva63@gmail.com",0,0.00,0.00),
        ("Shang","nshang946@gmail.com",0,0.00,0.00),
        ("Donn","odegaarddonald@gmail.com",0,0.00,0.00),
        ("Eva0","nduatieva85@gmail.com",1,0.00,100.00),
        ("Evans Ngotho","evansnthiwa04@gmail.com",0,0.00,0.00),
        ("Kagoto Lenny","kagotolenny8@gmail.com",1,0.00,100.00),
        ("Muneny","munenyian@gmail.com",1,0.00,100.00),
        ("Rmk","reubenqaris@gmail.com",0,0.00,0.00),
        ("GADDAFITERRY","logaddafi408@gmail.com",0,0.00,0.00),
        ("@kelvin","kelvinmutugi554@gmail.com",0,0.00,0.00),
        ("Renney","coldflame744@gmail.com",0,0.00,0.00),
        ("Colloh","iconclown@gmail.com",0,0.00,0.00),
        ("Xybe","catherinenjoroge513@gmail.com",0,0.00,0.00),
        ("notnice","ke795839@gmail.com",0,0.00,0.00),
        ("Blackburn","richardrichiemwe25@gmail.com",0,0.00,0.00),
        ("Lazar","lazarusmugo74@gmail.com",1,0.00,100.00),
        ("W1N5T00N","winstonrodney655@gmail.com",0,0.00,0.00),
        ("Mikz","mikzgitau55@gmail.com",0,0.00,0.00),
        ("Nick's","pmugambi975@gmail.com",0,0.00,0.00),
        ("Modgel","lpmodgel@gmail.com",0,0.00,0.00),
        ("Decks004","dekems004@gmail.com",0,0.00,0.00),
        ("Collo","collinskamanga514@gmail.com",0,0.00,0.00),
        ("Kimmy gal","njokililiana56@gmail.com",1,0.00,100.00),
        ("Incognito","derickeugine4@gmail.com",1,0.00,100.00),
        ("INEXUSTECH","mohdmwachiroho@gmail.com",0,0.00,0.00),
        ("George44","georgengaruiya03@gmail.com",0,0.00,0.00),
        ("Malack Kamanda","kamandamalack6@gmail.com",1,55.00,10.00),
        ("Colloraj","oirurucollins@gmail.com",0,0.00,0.00),
        ("Calx","vividelamanga@gmail.com",0,0.00,0.00),
        ("Rodgers","kilimorodgers108@gmail.com",0,0.00,0.00),
        ("Madollar","mwaswills852@gmail.com",0,0.00,0.00),
        ("Free soul","elvisonduko295@gmail.com",0,0.00,0.00),
        ("Telvin01","telvinkioko59@gmail.com",0,0.00,0.00),
        ("12asapray","njokipriyanka76@gmail.com",0,0.00,0.00),
        ("cinncy","khalibwacynthia@gmail.com",1,0.00,70.00),
        ("Paul","reconpaul02@gmail.com",0,0.00,0.00),
        ("enock4445","enockbett476@gmail.com",0,0.00,0.00),
        ("Govi","rangeralvine452@gmail.com",0,0.00,0.00),
        ("Alvine","alvinetiema@gmail.com",0,0.00,0.00),
        ("Simon","mwanziasimon38@gmail.com",1,0.00,100.00),
        ("cyrus","2504395@students.kcau.ac.ke",1,100.00,100.00),
        ("mack12","mackdaghwill4@gmail.com",0,0.00,0.00),
        ("Anthony","leyiananthony96@gmail.com",0,0.00,0.00),
        ("Rankii","mataramoses2007@gmail.com",1,0.00,100.00),
        ("Daniel","sifunadaniel3264@gmail.com",0,0.00,0.00),
        ("Big shock","grahamshock967@gmail.com",0,0.00,0.00),
        ("mallonte","mathengemarlon47@gmail.com",0,0.00,0.00),
        ("Mr. CEO","wanderabrian913@gmail.com",1,0.00,100.00),
        ("Lins","mabeyabrian73@gmail.com",0,0.00,0.00),
        ("Trent","enockkipkoech5220@gmail.com",1,0.00,100.00),
        ("FRESH-GANG","akamajohn06@gmail.com",1,0.00,100.00),
        ("Jffy","juliusmakenzie9@gmail.com",0,0.00,0.00),
        ("Julius","juliusmakenzie2004@gmail.com",0,0.00,0.00),
        ("Itskelvin","itskelvin2004@gmail.com",0,0.00,0.00),
        ("Edwin","edwinkanyi605@gmail.com",0,0.00,0.00),
        ("Ayanai","philiphkirui443@gmail.com",1,0.00,100.00),
        ("hebron","hebrontum24@gmail.com",0,0.00,0.00),
        ("Kamanu","thisiskamanu@gmail.com",0,0.00,0.00),
        ("Juma","wanyendedopher@gmail.com",1,0.00,100.00),
        ("Edward wuapari","edwardolewuapari@gmail.com",0,0.00,0.00),
        ("Meshtechand graphic","meshackouko627@gmail.com",1,0.00,40.00),
        ("Smyth","githaigdm002@gmail.com",0,0.00,0.00),
        ("Clement","clementkiprotich195@gmail.com",0,0.00,0.00),
        ("kukurela005","yassirkamau@gmail.com",0,0.00,0.00),
        ("Ronney","ronneycheb@gmail.com",0,0.00,0.00),
        ("CollinsKE","colepanther1@gmail.com",0,0.00,0.00),
        ("Aman6ix","fatumaabdallahsalim46@gmail.com",0,0.00,0.00),
        ("Buaz","miltonabaaz@gmail.com",0,0.00,0.00),
        ("Joseph","joseph24mwangi5@gmail.com",0,0.00,0.00),
        ("Bonyo","cosmasbonyo138@gmail.com",0,0.00,0.00),
        ("Samgift","sammykruze86@gmail.com",1,0.00,100.00),
        ("Nicolas","nicksavali946@gmail.com",0,0.00,0.00),
        ("kinaramacha39","kinaramacha39@gmail.com",1,0.00,140.00),
        ("Migos","marymigwi500@gmail.com",1,0.00,100.00),
        ("Abed28879","abedjared94@gmail.com",0,0.00,0.00),
        ("Bannyke","bannymavellah@gmail.com",0,0.00,0.00),
        ("Spoiler","kemboihezron119@gmail.com",1,0.00,100.00),
        ("Alekee","alexyegon158@gmail.com",0,0.00,0.00),
        ("MANDEM","alogytamathew@gmail.com",0,0.00,0.00),
        ("Kalebu","kipkoechngeno5209@gmail.com",0,0.00,0.00),
        ("Kings","stevehassan091@gmail.com",1,0.00,100.00),
        ("Null Sniffer","centralpopcee9@gmail.com",0,0.00,0.00),
        ("Zeddie","zeddiedickson@gmail.com",0,0.00,0.00),
        ("Klaus","mahigwaisaac18@gmail.com",0,0.00,0.00),
        ("Shelby Cyrus","cyrusrogers949@gmail.com",0,0.00,0.00),
        ("Ceaser","kamandejulius956@gmail.com",1,0.00,100.00),
        ("Bavoo","bavoo1239@gmail.com",1,0.00,100.00),
        ("Ralph254","ralphrhoderick@gmail.com",1,0.00,100.00),
        ("lammy kim","lammykim6872@gmail.com",1,0.00,100.00),
        ("ONDARA","obadiahondara257@gmail.com",0,0.00,0.00),
        ("duncantoo001","tooduncan871@gmail.com",1,0.00,100.00),
        ("daggy","coolkiddaggy@gmail.com",0,0.00,0.00),
        ("JAAYY.","nisgove@gmail.com",0,0.00,0.00),
        ("Infinite x","xhuncho732@gmail.com",0,0.00,0.00),
        ("keith504","cravenskeith29@gmail.com",0,0.00,0.00),
        ("Wenslaus","wenslauswanyonyi3@gmail.com",0,0.00,0.00),
        ("Eugraizz","opandaeugine@gmail.com",0,0.00,0.00),
        ("Only1d","derosd990@gmail.com",0,0.00,0.00),
        ("mwana fa","derrickderos982@gmail.com",0,0.00,0.00),
        ("GENZICT TECH","genzicttech@gmail.com",1,4300.00,29010.00),
        ("Faded simpson","ongereevans66@gmail.com",1,0.00,0.00),
        ("IRINE","milanoiirineirine@gmail.com",1,0.00,0.00),
        ("Matoo","sigeik477@gmail.com",1,0.00,0.00),
        ("samueleeugine","samueleugine166@gmail.com",1,0.00,0.00),
        ("Senior","abellimorono@gmail.com",1,0.00,0.00),
        ("Lupao wanyonyi","wanyonyialvin28@gmail.com",1,0.00,0.00),
        ("SAM'S TECH","sammy2wambua@gmail.com",1,0.00,0.00),
        ("Nicoh","nicosavaii5@gmail.com",1,0.00,0.00),
        ("Chumbaa","beatricechepchumba65@gmail.com",1,0.00,0.00),
        ("Pinchez004","iann03040@gmail.com",1,0.00,0.00),
        ("Vjay","videlis701@gmail.com",1,50.00,100.00),
        ("Brightbrin Richer","brightbrinricher@gmail.com",1,0.00,0.00),
        ("Tfx","langatgideon129@gmail.com",1,0.00,100.00),
        ("Noti","bildadbildad25@gmail.com",1,0.00,0.00),
        ("morriskarani","morriskaraniwawira@gmail.com",1,0.00,100.00),
        ("Judie","judiecherono9@gmail.com",1,0.00,0.00),
        ("Reagy","reagyke4@gmail.com",1,0.00,0.00),
        ("iiam.nashon","biannashon@gmail.com",1,0.00,0.00),
        ("Mroyal","ekimathi092@gmail.com",1,0.00,0.00),
        ("Tom","tokumu@gmail.com",1,0.00,0.00),
        ("Ushindi charo","randuchackso@gmail.com",1,0.00,100.00),
        ("pablo","pabloheroic10@gmail.com",1,0.00,0.00),
        ("centralpopcee","nullsniffer@gmail.com",1,0.00,0.00),
        ("Aleco","xelaaleco@gmail.com",1,0.00,0.00),
        ("Travis Elvis","traviselvis731@gmail.com",1,0.00,0.00),
        ("Felonyfest","flaakof@gmail.com",1,0.00,0.00),
        ("Ouma","lynnelexy976@gmail.com",1,0.00,0.00),
        ("KLKL","mackdRTRTRawill4@gmail.com",1,0.00,1.00),
        ("mackj","mackdawqwertyill4@gmail.com",0,10.00,0.00),
        ("tyty","tyty@gmail.com",1,0.00,0.50),
        ("tttt","test11@gmail.com",0,0.00,0.00),
        ("dhhdh","mackdgetxawill4@gmail.com",0,0.00,0.00),
        ("tyu","macrrrrkdawill4@gmail.com",0,0.00,0.00),
        ("merab167","mackdffffawill4@gmail.com",0,0.00,0.00),
        ("merab111","mackdsssawill4@gmail.com",1,0.00,0.50),
        ("merab11","mackdxxawill4@gmail.com",1,0.50,0.50),
    ]
    conn = get_db_connection()
    for username, email, is_active, wallet, trade in rows:
        email_lower = email.lower()
        existing = conn.execute("SELECT 1 FROM users WHERE LOWER(email)=?", (email_lower,)).fetchone()
        if not existing:
            conn.execute(
                """INSERT INTO users
                   (email, username, password_hash, phone, is_active, balance,
                    binary_balance, binary_deposited, referral_code, joined_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (email_lower, username, generate_password_hash("123456"),
                 "254700000000", is_active, wallet, trade, trade,
                 f"GP-{uuid.uuid4().hex.upper()[:5]}",
                 datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
            )
        else:
            conn.execute(
                """UPDATE users SET balance=?, binary_balance=?, binary_deposited=?,
                   is_active=?, username=? WHERE LOWER(email)=?""",
                (wallet, trade, trade, is_active, username, email_lower)
            )
    conn.commit()
    conn.close()


seed_active_users()


# ----------------------------------------------------------------
# AUTH DECORATOR
# ----------------------------------------------------------------
def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ----------------------------------------------------------------
# EMAIL HELPER — Resend
# ----------------------------------------------------------------
def send_reset_email(to_email: str, reset_link: str) -> bool:
    try:
        params = {
            "from": "GainPesa <onboarding@resend.dev>",
            "to": [to_email],
            "subject": "GainPesa – Password Reset Request",
            "html": f"""
                <div style="font-family:Arial,sans-serif;max-width:500px;margin:auto;padding:20px;">
                    <h2 style="color:#2e7d32;">GainPesa Password Reset</h2>
                    <p>You requested a password reset. Click the button below to set a new password:</p>
                    <a href="{reset_link}"
                       style="display:inline-block;background:#2e7d32;color:white;padding:12px 24px;
                              text-decoration:none;border-radius:6px;margin:16px 0;">
                        Reset My Password
                    </a>
                    <p style="color:#666;font-size:13px;">This link expires in <strong>1 hour</strong>.</p>
                    <p style="color:#666;font-size:13px;">If you did not request this, ignore this email.</p>
                    <hr style="border:none;border-top:1px solid #eee;margin-top:30px;">
                    <p style="color:#aaa;font-size:11px;">GainPesa &copy; {datetime.datetime.now().year}</p>
                </div>
            """,
        }
        response = resend.Emails.send(params)
        app.logger.info(f"[RESEND] Sent to {to_email}: {response}")
        return True
    except Exception as e:
        app.logger.error(f"[RESEND ERROR] {to_email}: {e}")
        # Console fallback — visible in Render logs
        print(f"\n{'='*65}")
        print(f"[RESET LINK — copy and open in browser]")
        print(f"To   : {to_email}")
        print(f"Link : {reset_link}")
        print(f"{'='*65}\n")
        return False


# ================================================================
# ROUTES
# ================================================================

@app.route("/")
def index():
    return render_template("index.html")


@app.route('/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')


@app.route('/sw.js')
def service_worker():
    return app.send_static_file('sw.js')


@app.route("/register", methods=["GET", "POST"])
def register():
    error = None
    ref_code = request.args.get("ref")
    if request.method == "POST":
        email       = request.form.get("email", "").strip().lower()
        username    = request.form.get("username", "").strip()
        password    = request.form.get("password", "")
        phone       = request.form.get("phone", "").strip()
        referred_by = request.form.get("ref", "").strip() or None
        conn = get_db_connection()
        if conn.execute("SELECT email FROM users WHERE LOWER(email)=?", (email,)).fetchone():
            error = "Email already exists"
        elif conn.execute("SELECT username FROM users WHERE username=?", (username,)).fetchone():
            error = "Username already taken"
        if error:
            conn.close()
            return render_template("register.html", error=error, ref_code=ref_code)
        conn.execute(
            "INSERT INTO users (email,username,password_hash,phone,referral_code,referred_by,joined_at) VALUES (?,?,?,?,?,?,?)",
            (email, username, generate_password_hash(password), phone,
             f"GP-{uuid.uuid4().hex.upper()[:5]}", referred_by,
             datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        )
        conn.commit()
        conn.close()
        session["user_email"] = email
        return redirect(url_for("pay_page"))
    return render_template("register.html", ref_code=ref_code)


@app.route("/login", methods=["GET", "POST"])
def login():
    error = request.args.get("error")
    if request.method == "POST":
        credential = request.form.get("credential", "").strip()
        password   = request.form.get("password", "")
        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE LOWER(email)=? OR username=?",
            (credential.lower(), credential)
        ).fetchone()
        conn.close()
        if not user or not check_password_hash(user["password_hash"], password):
            error = "Invalid credentials"
        else:
            session["user_email"] = user["email"]
            return redirect(url_for("dashboard") if user["is_active"] else url_for("pay_page"))
    return render_template("register.html", error=error)


# ================================================================
# PASSWORD RESET
# ================================================================

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        conn  = get_db_connection()
        try:
            user = conn.execute(
                "SELECT email FROM users WHERE LOWER(email)=?", (email,)
            ).fetchone()
            if user:
                token = serializer.dumps(user["email"], salt="gainpesa-password-reset")
                if os.getenv("RENDER"):
                    host       = os.getenv("RENDER_EXTERNAL_HOSTNAME", "gainpesa-zatz.onrender.com")
                    reset_link = f"https://{host}/reset-password/{token}"
                else:
                    reset_link = url_for("reset_password", token=token, _external=True)
                sent = send_reset_email(user["email"], reset_link)
                if sent:
                    flash("Reset link sent! Check your inbox (and spam folder).", "info")
                else:
                    flash(
                        "If that email is registered, check your inbox. "
                        "If no email arrives, contact support — the link was logged on the server.",
                        "info"
                    )
            else:
                flash("If that email is registered, a reset link has been sent.", "info")
        except Exception as e:
            app.logger.error(f"[ForgotPassword] Error: {e}")
            flash("Something went wrong. Please try again.", "error")
        finally:
            conn.close()
        return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    try:
        email = serializer.loads(token, salt="gainpesa-password-reset", max_age=3600)
    except SignatureExpired:
        flash("This reset link has expired (1 hour limit). Please request a new one.", "error")
        return redirect(url_for("forgot_password"))
    except (BadSignature, Exception):
        flash("This reset link is invalid or has already been used.", "error")
        return redirect(url_for("forgot_password"))

    conn = get_db_connection()
    user = conn.execute("SELECT email FROM users WHERE LOWER(email)=?", (email.lower(),)).fetchone()
    if not user:
        conn.close()
        flash("Account not found.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        new_pw  = request.form.get("password", "")
        conf_pw = request.form.get("confirm_password", "")
        if len(new_pw) < 6:
            conn.close()
            return render_template("reset_password.html", token=token,
                                   error="Password must be at least 6 characters.")
        if new_pw != conf_pw:
            conn.close()
            return render_template("reset_password.html", token=token,
                                   error="Passwords do not match.")
        try:
            conn.execute("UPDATE users SET password_hash=? WHERE LOWER(email)=?",
                         (generate_password_hash(new_pw), email.lower()))
            conn.commit()
        except Exception as e:
            app.logger.error(f"[ResetPassword] Update error: {e}")
            conn.close()
            return render_template("reset_password.html", token=token,
                                   error="Failed to save new password. Please try again.")
        conn.close()
        flash("✓ Password updated successfully! You can now log in.", "success")
        return redirect(url_for("login"))

    conn.close()
    return render_template("reset_password.html", token=token)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/pay")
@login_required
def pay_page():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email=?", (session["user_email"],)).fetchone()
    conn.close()
    # Already active — skip pay page
    if user and user["is_active"]:
        return redirect(url_for("dashboard"))
    return render_template("pay.html", user=dict(user) if user else {})


@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email=?", (session["user_email"],)).fetchone()
    conn.close()
    if not user["is_active"]:
        return redirect(url_for("pay_page"))
    return render_template("dashboard.html", user=dict(user))


@app.route("/gainbinary")
@login_required
def gainbinary():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email=?", (session["user_email"],)).fetchone()
    conn.close()
    if not user["is_active"]:
        return redirect(url_for("pay_page"))
    return render_template("gainbinary.html", user=dict(user))


# ================================================================
# PAYMENT: ACTIVATION
# ================================================================

@app.route("/api/initiate-payment", methods=["POST"])
@login_required
def initiate_payment():
    email = session["user_email"]
    conn  = get_db_connection()
    user  = conn.execute("SELECT phone, is_active FROM users WHERE email=?", (email,)).fetchone()
    conn.close()

    if user["is_active"]:
        return jsonify({"success": False, "error": "Account already active"}), 400

    phone = user["phone"] or ""
    if phone.startswith("0"):   phone = "254" + phone[1:]
    elif phone.startswith("+"): phone = phone[1:]

    ext_ref = "GP-ACT-" + uuid.uuid4().hex[:6].upper()
    try:
        r = requests.post(
            f"{PAYHERO_BASE_URL}/payments",
            json={
                "amount": ACTIVATION_FEE, "phone_number": phone,
                "channel_id": PAYHERO_CHANNEL_ID, "provider": PAYHERO_PROVIDER,
                "external_reference": ext_ref, "callback_url": CALLBACK_URL
            },
            headers={"Content-Type": "application/json", "Authorization": get_auth_header()},
            timeout=15
        )
        app.logger.info(f"[PAYHERO] status={r.status_code} body={r.text[:300]}")
        if r.status_code in [200, 201]:
            conn = get_db_connection()
            conn.execute(
                "INSERT INTO transactions (ext_ref,email,type,status,amount) VALUES (?,?,?,?,?)",
                (ext_ref, email, "activation", "pending", ACTIVATION_FEE)
            )
            conn.commit()
            conn.close()
            return jsonify({"success": True, "reference": ext_ref})
        return jsonify({"success": False, "error": r.text})
    except Exception as e:
        app.logger.error(f"[PAYHERO ERROR] {e}")
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/reconcile/<ext_ref>")
@login_required
def reconcile(ext_ref):
    conn = get_db_connection()
    tx   = conn.execute(
        "SELECT status FROM transactions WHERE ext_ref=? AND email=?",
        (ext_ref, session["user_email"])
    ).fetchone()
    conn.close()
    if not tx:
        return jsonify({"status": "not_found"}), 404
    if tx["status"] == "confirmed":
        return jsonify({"status": "confirmed"})
    elif tx["status"] == "failed":
        return jsonify({"status": "canceled"})
    else:
        return jsonify({"status": "pending"})


@app.route("/api/binary/deposit", methods=["POST"])
@login_required
def initiate_binary_deposit():
    amount = float(request.json.get("amount", 0))
    email  = session["user_email"]
    if amount < MIN_BINARY_DEPOSIT_KES:
        return jsonify({"error": f"Minimum deposit is Ksh {MIN_BINARY_DEPOSIT_KES:.0f} (~1 USD)"}), 400
    conn  = get_db_connection()
    phone = conn.execute("SELECT phone FROM users WHERE email=?", (email,)).fetchone()["phone"]
    conn.close()
    if phone.startswith("0"):   phone = "254" + phone[1:]
    elif phone.startswith("+"): phone = phone[1:]
    ext_ref = "GP-BIN-" + uuid.uuid4().hex[:6].upper()
    try:
        r = requests.post(
            f"{PAYHERO_BASE_URL}/payments",
            json={
                "amount": amount, "phone_number": phone,
                "channel_id": PAYHERO_CHANNEL_ID, "provider": PAYHERO_PROVIDER,
                "external_reference": ext_ref, "callback_url": CALLBACK_URL
            },
            headers={"Content-Type": "application/json", "Authorization": get_auth_header()},
            timeout=15
        )
        if r.status_code in [200, 201]:
            conn = get_db_connection()
            conn.execute(
                "INSERT INTO transactions (ext_ref,email,type,status,amount) VALUES (?,?,?,?,?)",
                (ext_ref, email, "binary_deposit", "pending", amount)
            )
            conn.commit()
            conn.close()
            return jsonify({"success": True, "reference": ext_ref})
        return jsonify({"success": False, "error": r.text})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/callback", methods=["POST"])
def callback():
    try:
        data      = request.json or {}
        res       = data.get("response") or data
        ext_ref   = res.get("ExternalReference")
        status    = str(res.get("Status", "")).lower()
        cb_amount = float(res.get("Amount", 0))

        app.logger.info(f"[CALLBACK] ref={ext_ref} status={status} amount={cb_amount}")

        if not ext_ref:
            return jsonify({"status": "error", "reason": "no ext_ref"}), 400

        conn = get_db_connection()
        tx   = conn.execute("SELECT * FROM transactions WHERE ext_ref=?", (ext_ref,)).fetchone()
        if not tx:
            conn.close()
            return jsonify({"status": "not_found"}), 404

        # Idempotency — already processed
        if tx["status"] == "confirmed":
            conn.close()
            return jsonify({"status": "ok", "note": "already confirmed"})

        if status not in ["success", "successful"]:
            conn.execute("UPDATE transactions SET status='failed' WHERE ext_ref=?", (ext_ref,))
            conn.commit()
            conn.close()
            return jsonify({"status": "ok"})

        tx_type   = tx["type"] or "activation"
        tx_amount = float(tx["amount"]) if tx["amount"] else cb_amount

        conn.execute("UPDATE transactions SET status='confirmed' WHERE ext_ref=?", (ext_ref,))

        if tx_type == "activation":
            conn.execute("UPDATE users SET is_active=1 WHERE email=?", (tx["email"],))
            conn.execute(
                "UPDATE users SET binary_balance=binary_balance+?, binary_deposited=binary_deposited+? WHERE email=?",
                (tx_amount, tx_amount, tx["email"])
            )
            # Referral commission: 50% to referrer
            ur = conn.execute("SELECT referred_by FROM users WHERE email=?", (tx["email"],)).fetchone()
            if ur and ur["referred_by"]:
                ref = conn.execute(
                    "SELECT email FROM users WHERE referral_code=?", (ur["referred_by"],)
                ).fetchone()
                if ref:
                    comm = round(tx_amount * 0.50, 2)
                    conn.execute(
                        "UPDATE users SET balance=balance+?, total_earned=total_earned+?, total_referred=total_referred+1 WHERE email=?",
                        (comm, comm, ref["email"])
                    )

        elif tx_type == "binary_deposit":
            conn.execute(
                "UPDATE users SET binary_balance=binary_balance+?, binary_deposited=binary_deposited+? WHERE email=?",
                (tx_amount, tx_amount, tx["email"])
            )

        conn.commit()
        conn.close()
        return jsonify({"status": "ok"})

    except Exception as e:
        app.logger.error(f"[CALLBACK ERROR] {e}", exc_info=True)
        return jsonify({"status": "error", "reason": str(e)}), 500


# ================================================================
# TRADING APIs
# ================================================================

@app.route("/api/binary/trade", methods=["POST"])
@login_required
def execute_binary_trade():
    data   = request.json
    email  = session["user_email"]
    amount = float(data.get("amount", 0))
    conn   = get_db_connection()
    user   = conn.execute("SELECT binary_balance FROM users WHERE email=?", (email,)).fetchone()
    if user["binary_balance"] < amount:
        conn.close()
        return jsonify({"error": "Insufficient Trading Balance"}), 400

    # 100% win — 80% profit every trade
    payout = round(amount * 1.8, 2)
    profit = round(amount * 0.8, 2)

    conn.execute(
        "UPDATE users SET binary_balance=binary_balance-?+? WHERE email=?",
        (amount, payout, email)
    )
    # binary_winnings = unclaimed profits; total_earned = all-time profit tracker
    conn.execute(
        "UPDATE users SET binary_winnings=binary_winnings+?, total_earned=total_earned+? WHERE email=?",
        (profit, profit, email)
    )
    conn.execute(
        "INSERT INTO binary_trades (email,asset,amount,direction,status,payout,timestamp) VALUES (?,?,?,?,?,?,?)",
        (email, data.get("asset", "EUR/USD"), amount, data.get("direction"), "win", payout,
         datetime.datetime.now().strftime("%H:%M:%S"))
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "status": "win", "payout": payout, "profit": profit})


@app.route("/api/binary/claim-winnings", methods=["POST"])
@login_required
def claim_binary_winnings():
    email  = session["user_email"]
    amount = float(request.json.get("amount", 0))
    conn   = get_db_connection()
    user   = conn.execute(
        "SELECT binary_winnings, binary_balance FROM users WHERE email=?", (email,)
    ).fetchone()
    if amount <= 0:
        conn.close()
        return jsonify({"error": "Invalid amount"}), 400
    if amount > round(user["binary_winnings"], 2):
        conn.close()
        return jsonify({"error": f"Available winnings: Ksh {user['binary_winnings']:.2f}. Deposited capital cannot be withdrawn."}), 400
    if amount > user["binary_balance"]:
        conn.close()
        return jsonify({"error": "Insufficient trading balance"}), 400
    # Move winnings → wallet (total_earned already incremented at trade time)
    conn.execute(
        "UPDATE users SET binary_balance=binary_balance-?, binary_winnings=binary_winnings-?, balance=balance+? WHERE email=?",
        (amount, amount, amount, email)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/binary/transfer", methods=["POST"])
@login_required
def transfer_to_binary():
    amount = float(request.json.get("amount", 0))
    email  = session["user_email"]
    conn   = get_db_connection()
    user   = conn.execute("SELECT balance FROM users WHERE email=?", (email,)).fetchone()
    if user["balance"] < amount:
        conn.close()
        return jsonify({"error": "Insufficient Wallet Balance"}), 400
    conn.execute(
        "UPDATE users SET balance=balance-?, binary_balance=binary_balance+?, binary_deposited=binary_deposited+? WHERE email=?",
        (amount, amount, amount, email)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ================================================================
# USER DATA API — real total_earned
# ================================================================

@app.route("/api/user", methods=["GET"])
@login_required
def get_user_data():
    email = session["user_email"]
    conn  = get_db_connection()
    user  = conn.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
    withdrawals = conn.execute(
        "SELECT amount, mpesa_number as mpesa, status, date FROM withdrawals WHERE email=? ORDER BY id DESC",
        (email,)
    ).fetchall()
    conn.close()

    # Real total earned:
    # total_earned (DB) = binary profits + referral commissions + admin wallet credits
    # binary_winnings   = unclaimed profits still sitting in trade balance
    # balance           = wallet (claimed winnings + referral comms available to withdraw)
    stored_earned  = float(user["total_earned"] or 0)
    unclaimed_wins = float(user["binary_winnings"] or 0)
    real_total     = round(stored_earned + unclaimed_wins, 2)

    return jsonify({
        "balance":             float(user["balance"] or 0),
        "binary_balance":      float(user["binary_balance"] or 0),
        "binary_deposited":    float(user["binary_deposited"] or 0),
        "binary_winnings":     float(user["binary_winnings"] or 0),
        "withdrawable_balance": float(user["balance"] or 0),
        "total_earned":        real_total,
        "total_withdrawn":     float(user["total_withdrawn"] or 0),
        "total_referred":      user["total_referred"],
        "referral_code":       user["referral_code"],
        "min_binary_deposit":  MIN_BINARY_DEPOSIT_KES,
        "withdrawals":         [dict(w) for w in withdrawals],
    })


@app.route("/api/withdraw", methods=["POST"])
@login_required
def submit_withdraw():
    email  = session["user_email"]
    amount = float(request.json.get("amount", 0))
    mpesa  = request.json.get("mpesa", "")
    if amount < 300:
        return jsonify({"error": "Minimum withdrawal is Ksh 300"}), 400
    conn  = get_db_connection()
    avail = round(float(
        conn.execute("SELECT balance FROM users WHERE email=?", (email,)).fetchone()["balance"] or 0
    ), 2)
    if amount > avail:
        conn.close()
        return jsonify({"error": f"Only your earnings can be withdrawn. Available: Ksh {avail:.2f}"}), 400
    conn.execute(
        "UPDATE users SET balance=balance-?, total_withdrawn=total_withdrawn+? WHERE email=?",
        (amount, amount, email)
    )
    conn.execute(
        "INSERT INTO withdrawals (email,amount,mpesa_number,status,date) VALUES (?,?,?,?,?)",
        (email, amount, mpesa, "pending", datetime.datetime.now().strftime("%b %d, %Y %H:%M"))
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ================================================================
# ADMIN
# ================================================================

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("username") == "MACK" and request.form.get("password") == "AJEGA":
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
    return render_template("admin_login.html")


@app.route("/admin")
def admin_dashboard():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    conn = get_db_connection()
    users = conn.execute("SELECT * FROM users ORDER BY joined_at DESC").fetchall()
    withdrawals = conn.execute(
        "SELECT w.*, u.username FROM withdrawals w JOIN users u ON w.email=u.email ORDER BY w.id DESC"
    ).fetchall()
    recent_updates = conn.execute(
        "SELECT l.*, u.username FROM admin_logs l JOIN users u ON l.target_email=u.email ORDER BY l.id DESC LIMIT 30"
    ).fetchall()
    conn.close()
    return render_template("admin.html",
                           users=[dict(u) for u in users],
                           withdrawals=[dict(w) for w in withdrawals],
                           recent_updates=[dict(r) for r in recent_updates])


@app.route("/admin/update-balance", methods=["POST"])
def admin_update_balance():
    if not session.get("is_admin"):
        return jsonify({"error": "Unauthorized"}), 403
    email = request.json.get("email")
    amt   = float(request.json.get("balance", 0))
    conn  = get_db_connection()
    conn.execute(
        "UPDATE users SET balance=balance+?, total_earned=total_earned+? WHERE email=?",
        (amt, amt, email)
    )
    conn.execute(
        "INSERT INTO admin_logs (admin_username,target_email,action_type,amount,timestamp) VALUES (?,?,?,?,?)",
        ("MACK", email, "Wallet Addition", amt, datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/admin/update-trading", methods=["POST"])
def admin_update_trading():
    if not session.get("is_admin"):
        return jsonify({"error": "Unauthorized"}), 403
    email = request.json.get("email")
    amt   = float(request.json.get("amount", 0))
    conn  = get_db_connection()
    conn.execute("UPDATE users SET binary_balance=binary_balance+? WHERE email=?", (amt, email))
    conn.execute(
        "INSERT INTO admin_logs (admin_username,target_email,action_type,amount,timestamp) VALUES (?,?,?,?,?)",
        ("MACK", email, "Binary Addition", amt, datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/admin/mark-paid", methods=["POST"])
def admin_mark_paid():
    if not session.get("is_admin"):
        return jsonify({"error": "Unauthorized"}), 403
    conn = get_db_connection()
    conn.execute("UPDATE withdrawals SET status='paid' WHERE id=?", (request.json.get("id"),))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/admin/activate-user", methods=["POST"])
def admin_activate_user():
    if not session.get("is_admin"):
        return jsonify({"error": "Unauthorized"}), 403
    email = request.json.get("email")
    conn  = get_db_connection()
    conn.execute("UPDATE users SET is_active=1 WHERE email=?", (email,))
    conn.execute(
        "UPDATE users SET binary_balance=binary_balance+?, binary_deposited=binary_deposited+? WHERE email=?",
        (ACTIVATION_FEE, ACTIVATION_FEE, email)
    )
    ur = conn.execute("SELECT referred_by FROM users WHERE email=?", (email,)).fetchone()
    if ur and ur["referred_by"]:
        ref = conn.execute(
            "SELECT email FROM users WHERE referral_code=?", (ur["referred_by"],)
        ).fetchone()
        if ref:
            comm = round(ACTIVATION_FEE * 0.50, 2)
            conn.execute(
                "UPDATE users SET balance=balance+?, total_earned=total_earned+?, total_referred=total_referred+1 WHERE email=?",
                (comm, comm, ref["email"])
            )
    conn.execute(
        "INSERT INTO admin_logs (admin_username,target_email,action_type,amount,timestamp) VALUES (?,?,?,?,?)",
        ("MACK", email, "Manual Activation", ACTIVATION_FEE, datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/admin/download-pdf/<status>")
def download_users_pdf(status):
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    conn = get_db_connection()
    if status == "activated":
        users = conn.execute("SELECT * FROM users WHERE is_active=1").fetchall()
    elif status == "pending":
        users = conn.execute("SELECT * FROM users WHERE is_active=0").fetchall()
    else:
        users = conn.execute("SELECT * FROM users").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{status.title()} Users"

    green  = "1B5E20"
    lgreen = "C8E6C9"
    white  = "FFFFFF"
    grey   = "F5F5F5"

    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value     = f"GAINPESA — {status.upper()} USERS REPORT"
    tc.font      = Font(name="Arial", bold=True, size=14, color=white)
    tc.fill      = PatternFill("solid", start_color=green)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    dc = ws["A2"]
    dc.value     = f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
    dc.font      = Font(name="Arial", italic=True, size=9, color="555555")
    dc.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 16

    headers = ["#", "Email", "Username", "Phone", "Wallet (Ksh)", "Status", "Joined At"]
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=white)
        cell.fill      = PatternFill("solid", start_color=green)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[3].height = 20

    for ri, u in enumerate(users, start=1):
        er = ri + 3
        fc = lgreen if ri % 2 == 0 else grey
        row_data = [
            ri, u["email"], u["username"], u["phone"] or "",
            round(float(u["balance"] or 0), 2),
            "Active" if u["is_active"] else "Pending",
            u["joined_at"] or "",
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", start_color=fc)
            cell.alignment = Alignment(
                horizontal="center" if ci in [1, 5, 6] else "left",
                vertical="center"
            )
            cell.border = border
            if ci == 5:
                cell.number_format = '#,##0.00'
            if ci == 6:
                cell.font = Font(name="Arial", size=9,
                                 color="1B5E20" if u["is_active"] else "B71C1C", bold=True)

    total_row = len(users) + 4
    ws.cell(row=total_row, column=4, value="TOTAL WALLET").font = Font(bold=True, name="Arial", size=9)
    tc2 = ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row - 1})")
    tc2.font          = Font(bold=True, name="Arial", size=9, color=green)
    tc2.number_format = '#,##0.00'
    tc2.border        = border

    col_widths = [5, 35, 18, 18, 16, 12, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.read())
    resp.headers.set("Content-Disposition", "attachment", filename=f"{status}_users.xlsx")
    resp.headers.set("Content-Type",
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return resp


# ================================================================
# DEBUG
# ================================================================

@app.route("/debug-mail")
def debug_mail():
    return jsonify({
        "RESEND_API_KEY_SET": bool(os.getenv("RESEND_API_KEY")),
        "RENDER": os.getenv("RENDER"),
        "CALLBACK_URL": CALLBACK_URL,
    })


@app.route("/test-mail")
def test_mail():
    try:
        params = {
            "from": "GainPesa <onboarding@resend.dev>",
            "to": ["delivered@resend.dev"],
            "subject": "GainPesa Test Email",
            "text": "Resend is working correctly on Render.",
        }
        response = resend.Emails.send(params)
        return jsonify({"success": True, "response": str(response)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


if __name__ == "__main__":
    app.run(debug=True)
